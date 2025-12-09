import streamlit as st
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

# ==================================================
# STREAMLIT APP SETTINGS

st.set_page_config(page_title="Credit Risk Predictor", layout="wide")

# Title and short description
st.title("🔍 Credit Card Early Risk Detection System")
st.write("Upload CSV/XLSX → Get Risk Analysis → Download PDF & Excel outputs.")


# REQUIRED COLUMNS EXPECTED IN THE FILE

required_cols = [
    "Credit_Limit",
    "Utilisation_%",
    "DPD",
    "Monthly_Spend_Change",
    "Min_Due_Flag",
    "Merchant_Mix_Index",
    "Cash_Withdrawal_Frequency"
]

# FILE UPLOAD COMPONENT

uploaded_file = st.file_uploader("Upload CSV or Excel File", type=["csv", "xlsx"])

if uploaded_file:
    name = uploaded_file.name.lower()

    # Read CSV or Excel
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
        st.success("CSV loaded successfully!")
    else:
        df = pd.read_excel(uploaded_file)
        st.success("Excel file loaded successfully!")

    # Show preview
    st.write("📄 **Data Preview**")
    st.dataframe(df.head())

    # CHECK WHETHER ALL REQUIRED COLUMNS EXIST

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"Missing columns: {missing}")
        st.stop()


    # AUTO-GENERATE TARGET COLUMN IF NOT PRESENT

    if "Delinquent_30Plus" not in df.columns:

        # Rule-based conditions to estimate delinquency
        df["Delinquent_30Plus"] = (
            (df["Utilisation_%"] > 80).astype(int) +
            (df["Min_Due_Flag"] == 1).astype(int) +
            (df["DPD"] >= 15).astype(int) +
            (df["Monthly_Spend_Change"] < -0.4).astype(int) +
            (df["Merchant_Mix_Index"] < 0.3).astype(int)
        )

        # Final binary target (>= 2 conditions)
        df["Delinquent_30Plus"] = (df["Delinquent_30Plus"] >= 2).astype(int)

    # ==================================================
    # MACHINE LEARNING MODEL TRAINING
    # ==================================================
    X = df[required_cols]
    y = df["Delinquent_30Plus"]

    # Scale the features
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # Random forest model
    model = RandomForestClassifier(
        n_estimators=200,
        max_depth=5,
        random_state=42,
        class_weight="balanced"
    )
    model.fit(X_scaled, y)

    # Model outputs
    df["Risk_Probability"] = model.predict_proba(X_scaled)[:, 1]

    # Convert probability → labels
    df["Risk_Label"] = pd.cut(
        df["Risk_Probability"],
        bins=[0, 0.33, 0.66, 1],
        labels=["Low Risk", "Medium Risk", "High Risk"]
    )

    # SUMMARY SECTION

    st.subheader("📊 Portfolio Summary")

    total_customers = len(df)
    high_risk_pct = (df["Risk_Label"] == "High Risk").mean() * 100
    avg_util = df["Utilisation_%"].mean()
    avg_dpd = df["DPD"].mean()

    # Dashboard metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Customers", total_customers)
    c2.metric("High Risk %", f"{high_risk_pct:.2f}%")
    c3.metric("Avg Utilisation %", f"{avg_util:.2f}")
    c4.metric("Avg DPD", f"{avg_dpd:.2f}")

    # FLAG GENERATION (Behavioral)

    df["High_Utilisation_Flag"] = (df["Utilisation_%"] > 80).astype(int)
    df["Min_Due_Behavior_Flag"] = (df["Min_Due_Flag"] == 1).astype(int)
    df["Spend_Stress_Flag"] = (df["Monthly_Spend_Change"] < -0.4).astype(int)
    df["Essential_Spend_Spike_Flag"] = (df["Merchant_Mix_Index"] < 0.3).astype(int)
    df["DPD_Early_Warning"] = (df["DPD"] == 15).astype(int)
    df["Roll_Rate_Risk"] = (df["DPD"] == 30).astype(int)
    df["Severe_Roll_Risk"] = (df["DPD"] == 60).astype(int)


    # RECOMMENDED ACTIONS BASED ON RISK

    def action_map(risk):
        if risk == "Low Risk":
            return "👍 No action needed"
        if risk == "Medium Risk":
            return "⚠ Send reminder and encourage full payment"
        if risk == "High Risk":
            return "🚨 Proactive outreach: Offer payment plan/EMI options"
        return "Unknown"

    df["Recommended_Action"] = df["Risk_Label"].apply(action_map)

    st.subheader("🎯 Risk + Flags + Actions Table")
    st.dataframe(df.head())


    # VISUALIZATION (GAUGE + PIE)

    avg_risk = df["Risk_Probability"].mean()

    # Portfolio gauge
    g = go.Figure(go.Indicator(
        mode="gauge+number",
        value=avg_risk,
        title={"text": "Portfolio Risk Score"},
        gauge={
            "axis": {"range": [0, 1]},
            "steps": [
                {"range": [0, 0.33], "color": "green"},
                {"range": [0.33, 0.66], "color": "yellow"},
                {"range": [0.66, 1], "color": "red"},
            ],
        },
    ))
    st.plotly_chart(g, use_container_width=True)

    # Pie chart for distribution
    pie = px.pie(
        df,
        names="Risk_Label",
        title="Risk Distribution",
        color="Risk_Label",
        color_discrete_map={"Low Risk": "green", "Medium Risk": "yellow", "High Risk": "red"}
    )
    st.plotly_chart(pie, use_container_width=True)


    # EXCEL EXPORT WITH COLOR CODING

    output_excel = "Risk_Report_Colored.xlsx"
    df.to_excel(output_excel, index=False)

    wb = load_workbook(output_excel)
    ws = wb.active

    # Color formats
    red = PatternFill(start_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", fill_type="solid")

    # Risk column index
    risk_col_idx = list(df.columns).index("Risk_Label") + 1

    # Apply row color based on risk
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        risk = row[risk_col_idx - 1].value

        if risk == "High Risk":
            color = red
        elif risk == "Medium Risk":
            color = yellow
        else:
            color = green

        for cell in row:
            cell.fill = color

    wb.save(output_excel)

    # Download button
    with open(output_excel, "rb") as f:
        st.download_button("⬇ Download Colored Excel Report", f, file_name=output_excel)


    # PDF GENERATION FOR EACH CUSTOMER

    def generate_customer_pdf(row):
        cid = str(row["Customer_ID"])
        fname = f"Risk_Report_{cid}.pdf"

        c = canvas.Canvas(fname, pagesize=letter)
        w, h = letter

        # Title
        c.setFont("Helvetica-Bold", 18)
        c.drawString(50, h - 50, f"Customer Risk Report: {cid}")

        # Basic info
        c.setFont("Helvetica", 12)
        c.drawString(50, h - 90, "📌 Basic Information")
        c.drawString(70, h - 120, f"Credit Limit: {row['Credit_Limit']}")
        c.drawString(70, h - 140, f"Utilisation %: {row['Utilisation_%']}")
        c.drawString(70, h - 160, f"DPD: {row['DPD']} days")

        # Risk section
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, h - 200, "🎯 Risk Assessment")

        # Color based on risk
        if row["Risk_Label"] == "High Risk":
            c.setFillColor(colors.red)
        elif row["Risk_Label"] == "Medium Risk":
            c.setFillColor(colors.orange)
        else:
            c.setFillColor(colors.green)

        c.drawString(70, h - 230, f"Risk Level: {row['Risk_Label']}")
        c.setFillColor(colors.black)
        c.drawString(70, h - 250, f"Risk Probability: {round(row['Risk_Probability'], 3)}")

        # Flags section
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, h - 290, "🚩 Behavioural Flags")

        c.setFont("Helvetica", 12)
        y_pos = h - 320
        for col in row.index:
            if "Flag" in col and row[col] == 1:
                c.drawString(70, y_pos, f"- {col.replace('_', ' ')}")
                y_pos -= 20

        # Recommended Action
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y_pos - 20, "📝 Recommended Action")
        c.setFont("Helvetica", 12)
        c.drawString(70, y_pos - 50, row["Recommended_Action"])

        c.save()
        return fname


    # PDF DOWNLOAD SECTION

    st.subheader("📄 Customer-Level PDF Reports")

    customer_list = df["Customer_ID"].astype(str).tolist()
    selected_id = st.selectbox("Select Customer ID:", customer_list)

    if st.button("Generate PDF Report"):
        row = df[df["Customer_ID"].astype(str) == selected_id].iloc[0]
        pdf_file = generate_customer_pdf(row)

        with open(pdf_file, "rb") as f:
            st.download_button("⬇ Download PDF Report", f, file_name=pdf_file, mime="application/pdf")

        st.success("PDF Report Generated!")
